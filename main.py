import cv2
import os
import numpy as np
from sklearn.decomposition import PCA
import matplotlib.pyplot as plt
from openpyxl import *

directory = os.fsencode("Data set\Training")
wb = load_workbook(r"results.xlsx")
rowE = 1
rowM = 1
rowT = 1

listPR = []
#This function iterates through the images in (argument directory)
#convert each image to greyscale and resize it to 50*50
#then save the result in directory "Data set\\Trainpre\\"
def preprocessing(directory):
    imgNum = 1
    for filename in os.listdir(directory):
        fil = str(os.path.join(directory, filename).decode("utf-8"))
        image = cv2.imread(fil)
        gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        gray_image = cv2.resize(gray_image,(50,50));
        filename = filename.decode("utf-8")
        PCAfeatureExtraction(gray_image,imgNum)
        imgNum +=1
        #cv2.imwrite("Data set\\Trainpre\\"+ filename,gray_image)

def isolateObjects(imgPath):
    # Minimum percentage of pixels of same hue to consider dominant colour
    MIN_PIXEL_CNT_PCT = (1.0/20.0)
    image = cv2.imread(imgPath)
    image_hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
    h,_,_ = cv2.split(image_hsv)
    bins = np.bincount(h.flatten())
    peaks = np.where(bins > (h.size * MIN_PIXEL_CNT_PCT))[0]
    for i,peak in enumerate(peaks):
        mask = cv2.inRange(h, int(peak), int(peak))
        blob = cv2.bitwise_and(image, image, mask=mask)
        cv2.imwrite("Data set\\Testpre\\"+imgPath+"colourblobs-%d-hue_%03d.png" % (i, peak), blob)

def PCAfeatureExtraction(image,num):
    global rowM
    global rowE
    global rowT
    global listPR

    X = np.array(image)
    pca = PCA(n_components=2)
    pca.fit(X)
    #print(pca.explained_variance_ratio_)  
    #print(pca.singular_values_)
    meanVal = X.mean(axis=1)[:, np.newaxis]
    centered_matrix = X - meanVal #Subtract the Mean
    sheetTabM = wb['trainingimagesmean']
    for i in range(1,len(meanVal)):
        sheetTabM.cell(row=rowM, column=1).value = meanVal[i][0]
        rowM += 1     
    cov = np.dot(centered_matrix, centered_matrix.T) #Covariance matrix
    eigvals, eigvecs = np.linalg.eig(cov) #Calculate the Eigenvectors and Eigenvalues of the covariance matrix
    sheetTab = wb['eigvecs']   
    for i in range(1,len(eigvecs),1):
        for j in range(1,len(eigvecs[i]),1):
            sheetTab.cell(row=rowE, column=j).value = eigvecs[i][j]
        rowE += 1
    #Order eigenvectors by eigenvalues, highest to lowest.
    #Choose only the highest P eigenvectors. 
    # Make a list of (eigenvalue, eigenvector) tuples
    eig_pairs = [(np.abs(eigvals[i]), eigvecs[:,i]) for i in range(len(eigvals))]
    # Sort the (eigenvalue, eigenvector) tuples from high to low
    eig_pairs.sort()
    eig_pairs.reverse()
    # Visually confirm that the list is correctly sorted by decreasing eigenvalues
    #print('Eigenvalues in descending order:')
    
    for i in eig_pairs:
        listPR.append(i[0])
        #print()

preprocessing(directory)
sheetTabT = wb['training']

k = 0
for i in range(1,26):
    for j in range (2,20):
            sheetTabT.cell(row=i, column=1).value = int(i/25)+1
            sheetTabT.cell(row=i, column=j).value = listPR[k]
            k+=1  
    
wb.save("results.xlsx")